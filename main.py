import json
import os
from login_canaime import Login
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import math
import time

mostrar_navegador = False

url_pesquisa = (
    'https://canaime.com.br/sgp2rr/areas/unidades/pesquisa_resultadoVULGO.php?'
    'busca1=nome&busca2=&busca3=SAIDA&Submit2=Pesquisar'
)

def lista_ids_saida(nome_arquivo: str = "lista_ids_saida.json") -> list:
    try:
        page = Login(test=mostrar_navegador)
        page.goto(url_pesquisa)
    except Exception as e:
        print("Erro ao acessar a página de pesquisa:", e)
        return []

    try:
        qtd_presos_replace = ' REEDUCANDO(S) CADASTRADO(S)'
        qtd_presos_text = page.locator('.tituloVermelho10').text_content()
        qtd_presos = int(qtd_presos_text.replace(qtd_presos_replace, '').strip())
    except Exception as e:
        print("Erro ao extrair a quantidade de presos:", e)
        return []

    total_paginas = math.ceil(qtd_presos / 10)
    resultados = []

    pagina = 0
    tempos = []

    while pagina < total_paginas:
        inicio = time.time()
        paginas_restantes = total_paginas - (pagina + 1)
        print(f"Acessando página {pagina + 1} de {total_paginas}, faltam {paginas_restantes} páginas...")

        url_inicial = (
            f'https://canaime.com.br/sgp2rr/areas/unidades/pesquisa_resultadoVULGO.php?pageNum_rsPreso={pagina}'
            f'&totalRows_rsPreso={qtd_presos}&busca1=nome&busca2=&busca3=SAIDA&Submit2=Pesquisar'
        )

        try:
            page.goto(url_inicial)
            elementos = page.locator('a.tituloAzul')
            total_elementos = elementos.count()
        except Exception as e:
            print("Erro ao acessar a página ou localizar elementos:", e)
            break

        for i in range(total_elementos):
            try:
                elemento = elementos.nth(i)
                href = elemento.get_attribute('href')
                if href and "id_cad_preso=" in href:
                    id_cad_preso = href.split('=')[-1]
                else:
                    id_cad_preso = None
                nome = elemento.text_content().strip()
                resultados.append({'id': id_cad_preso, 'nome': nome})
            except Exception as e:
                print(f"Erro ao processar elemento {i} da página {pagina + 1}:", e)

        pagina += 1
        fim = time.time()
        tempos.append(fim - inicio)
        if len(tempos) > 0 and paginas_restantes > 0:
            tempo_medio = sum(tempos) / len(tempos)
            tempo_estimado = tempo_medio * paginas_restantes
            estimativa_timedelta = timedelta(seconds=tempo_estimado)
            horas, resto = divmod(estimativa_timedelta.seconds, 3600)
            minutos, segundos = divmod(resto, 60)
            print(f"Tempo estimado restante: {horas} horas, {minutos} minutos e {segundos} segundos.")
            print()

    try:
        with open(nome_arquivo, 'w', encoding='utf-8') as f:
            json.dump(resultados, f, ensure_ascii=False, indent=4)
        print(f"Lista de IDs salva no arquivo {nome_arquivo}")
    except Exception as e:
        print("Erro ao salvar o arquivo JSON:", e)

    return [page, resultados]


def busca_dados(lista_arquivo: str = "lista_ids_saida.json") -> list:
    if not os.path.exists(lista_arquivo) or os.path.getsize(lista_arquivo) == 0:
        print(f"Arquivo {lista_arquivo} não encontrado ou vazio.")
        return []
    else:
        print(f"Arquivo {lista_arquivo} encontrado. Carregando lista de IDs...")
        try:
            with open(lista_arquivo, 'r', encoding='utf-8') as f:
                lista_de_ids = json.load(f)
            page = Login(test=mostrar_navegador)
            return [page, lista_de_ids]
        except Exception as e:
            print("Erro ao carregar o arquivo JSON:", e)
            return []


def busca_datas(lista: list) -> list:
    try:
        page, lista = lista
    except ValueError:
        print("Lista de IDs inválida.")
        return []

    lista_presos_saida = []
    tempos = []
    url_certidao = "https://canaime.com.br/sgp2rr/areas/impressoes/UND_CertidaoCarceraria.php?id_cad_preso="

    for index, item in enumerate(lista):
        inicio = time.time()
        total_items = len(lista)
        items_restantes = total_items - (index + 1)
        print(f"Acessando preso {index + 1} de {total_items}, faltam {items_restantes} presos...")

        # Tentativa repetitiva em caso de timeout
        while True:
            try:
                page.goto(url_certidao + item['id'])
                lista_unit = page.locator('table+ table td.titulobk:nth-child(1)').all_text_contents()
                lista_datas = page.locator('table+ table .titulobk+ .titulobk:nth-child(2)').all_text_contents()

                for index_unit in reversed(range(len(lista_unit))):
                    if lista_unit[index_unit] not in ("SAIDA", "SAÍDA", "SA�DA"):
                        ultima_unit = lista_unit[index_unit].strip()
                        ultima_data = lista_datas[index_unit].strip()
                        print(item['id'], item['nome'], ultima_unit, ultima_data)
                        try:
                            data_convertida = datetime.strptime(ultima_data, "%d/%m/%Y")
                            if data_convertida.year == 2024:
                                lista_presos_saida.append({
                                    'Código': item['id'],
                                    'Preso': item['nome'],
                                    'Unidade': ultima_unit,
                                    'Data': ultima_data
                                })
                        except ValueError:
                            lista_presos_saida.append({
                                'Código': item['id'],
                                'Preso': item['nome'],
                                'Unidade': ultima_unit,
                                'Data': ultima_data
                            })
                        break
                # Se chegou até aqui, foi bem sucedido, então saia do loop infinito
                break

            except TimeoutError as te:
                print(f"Timeout ao acessar o preso {item.get('nome', '')} (ID: {item.get('id', '')}). Tentando novamente...")
                # Continua no loop, tentando novamente
            except Exception as e:
                # Outros erros não relacionados a timeout também podem ser tratados.
                # Se quiser tentar infinitamente só para TimeoutError, pode diferenciar aqui.
                print(f"Erro ao processar dados do preso {item.get('nome', '')} (ID: {item.get('id', '')}): {e}")
                # Dependendo do caso, se quiser tentar novamente apenas em caso de TimeoutError, pode colocar um break aqui para não loopar eternamente outros erros.
                # break

        fim = time.time()
        tempos.append(fim - inicio)

        if len(tempos) > 0 and items_restantes > 0:
            tempo_medio = sum(tempos) / len(tempos)
            tempo_estimado = tempo_medio * items_restantes
            estimativa_timedelta = timedelta(seconds=tempo_estimado)
            horas, resto = divmod(estimativa_timedelta.seconds, 3600)
            minutos, segundos = divmod(resto, 60)
            print(f"Tempo estimado restante: {horas} horas, {minutos} minutos e {segundos} segundos.")
            print()

    nome_arquivo_json = "resultado.json"
    try:
        with open(nome_arquivo_json, 'w', encoding='utf-8') as f:
            json.dump(lista_presos_saida, f, ensure_ascii=False, indent=4)
        print(f"Dados salvos em {nome_arquivo_json}")
    except Exception as e:
        print("Erro ao salvar o arquivo JSON:", e)

    return lista_presos_saida


def salvar_excel(lista_presos_saida: list, nome_arquivo: str = "presos_saida.xlsx") -> None:
    if not lista_presos_saida:
        print("A lista de presos está vazia, nada a salvar.")
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presos Saída"

        cabecalhos = ["Código", "Preso", "Unidade", "Data"]
        ws.append(cabecalhos)

        base_url = "https://canaime.com.br/sgp2rr/areas/unidades/Ficha_Menu.php?id_cad_preso="

        for preso in lista_presos_saida:
            codigo = preso.get("Código", "")
            preso_nome = preso.get("Preso", "")
            unidade = preso.get("Unidade", "")
            data = preso.get("Data", "")

            nova_linha = ws.max_row + 1
            ws.cell(row=nova_linha, column=2, value=preso_nome)
            ws.cell(row=nova_linha, column=3, value=unidade)
            ws.cell(row=nova_linha, column=4, value=data)

            cell_codigo = ws.cell(row=nova_linha, column=1, value=codigo)
            if codigo:
                cell_codigo.hyperlink = base_url + str(codigo)
                cell_codigo.style = "Hyperlink"

        for col_num, col_title in enumerate(cabecalhos, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(len(col_title), 20)

        wb.save(nome_arquivo)
        print(f"Arquivo Excel salvo como {nome_arquivo}")
    except Exception as e:
        print("Erro ao salvar o arquivo Excel:", e)


def main():
    # Caminho 1: Se existir resultado.json, então salvar excel.
    if os.path.isfile("resultado.json") and os.path.getsize("resultado.json") > 0:
        with open("resultado.json", "r", encoding="utf-8") as f:
            dados = json.load(f)
        if dados:
            salvar_excel(dados, "presos_saida.xlsx")
            return
        else:
            print("resultado.json está vazio.")

    # Caminho 2: Se não existir resultado.json, verifica se existe lista_ids_saida.json
    if os.path.isfile("lista_ids_saida.json") and os.path.getsize("lista_ids_saida.json") > 0:
        dados_ids = busca_dados("lista_ids_saida.json")  # Carrega a lista de ids
        if dados_ids and len(dados_ids) == 2:
            # Agora gera o resultado.json
            dados_presos = busca_datas(dados_ids)
            if dados_presos:
                salvar_excel(dados_presos, "presos_saida.xlsx")
                return
            else:
                print("Não foi possível obter dados detalhados a partir de lista_ids_saida.json.")
        else:
            print("Não foi possível carregar a lista de IDs a partir de lista_ids_saida.json.")
    else:
        # Caminho 3: Se não existir lista_ids_saida.json, então gera
        dados_ids = lista_ids_saida("lista_ids_saida.json")
        if dados_ids and len(dados_ids) == 2:
            dados_presos = busca_datas(dados_ids)
            if dados_presos:
                salvar_excel(dados_presos, "presos_saida.xlsx")
                return
            else:
                print("Não foi possível obter dados detalhados a partir da lista recém-criada.")
        else:
            print("Não foi possível obter a lista de IDs.")

if __name__ == "__main__":
    main()
